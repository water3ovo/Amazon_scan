from __future__ import annotations

from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ScanResult, ScanTarget
from .utils import clean_text, extract_asin

ACTIVE_VALUES = {"在投", "投放中", "active", "yes", "y", "true", "1"}

HEADER_ALIASES = {
    "国家": ["国家", "country", "Country"],
    "类型": ["类型", "type", "Type"],
    "Portfolio/品牌": ["Portfolio/品牌", "Portfolio", "品牌", "portfolio", "brand"],
    "产品": ["产品", "型号", "product", "Product"],
    "ASIN": ["ASIN", "asin"],
    "配置": ["配置", "configuration", "config"],
    "颜色": ["颜色", "color", "colour"],
    "URL": ["URL", "url", "链接", "商品链接"],
    "在投状态": ["在投状态", "状态", "active_status", "status"],
    "备注": ["备注", "remark", "notes"],
}

OWN_HEADERS = [
    "日期", "国家", "Portfolio", "产品", "ASIN", "URL", "产品名称", "配置", "颜色",
    "价格原始", "价格数值", "划线价格", "购买框归属", "库存预警", "配送信息", "配送>10天",
    "评分", "评论数", "BSR一级", "BSR二级", "异常标注", "优惠标签", "Amazon精选", "备注",
    "购买框状态",
]
COMP_HEADERS = [
    "日期", "国家", "品牌", "产品", "ASIN", "URL", "产品名称", "配置", "颜色",
    "价格原始", "价格数值", "划线价格", "购买框归属", "库存预警", "配送信息", "配送>10天",
    "评分", "评论数", "BSR一级", "BSR二级",
]
TECH_HEADERS = [
    "扫描时间", "国家", "类型", "ASIN", "URL", "scan_status", "error_reason", "attempts",
    "page_state", "stock_status", "购买框状态", "购买框归属", "配送方", "页面标题", "debug_screenshot",
]


def _header_lookup(headers: list[str]) -> dict[str, int]:
    normalized = {clean_text(h): i for i, h in enumerate(headers) if clean_text(h)}
    result = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[canonical] = normalized[alias]
                break
    return result


def _value(row, lookup: dict[str, int], key: str) -> str:
    idx = lookup.get(key)
    if idx is None or idx >= len(row):
        return ""
    return clean_text(row[idx])


def load_targets(path: str | Path, filter_active_products: bool = True) -> list[ScanTarget]:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"输入文件不存在: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    if "scan_targets" in wb.sheetnames:
        ws = wb["scan_targets"]
    elif "Mapping" in wb.sheetnames:
        ws = wb["Mapping"]
    else:
        ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [clean_text(x) for x in next(rows)]
    except StopIteration:
        return []
    lookup = _header_lookup(headers)
    if "国家" not in lookup:
        raise ValueError("输入表缺少‘国家’列。")
    if "ASIN" not in lookup and "URL" not in lookup:
        raise ValueError("输入表至少需要‘ASIN’或‘URL’列。")

    merged: OrderedDict[tuple[str, str], ScanTarget] = OrderedDict()
    for row in rows:
        country = _value(row, lookup, "国家").upper()
        url = _value(row, lookup, "URL")
        asin = _value(row, lookup, "ASIN").upper() or extract_asin(url)
        if not country or not asin:
            continue
        product_type = _value(row, lookup, "类型") or "本品"
        active_status = _value(row, lookup, "在投状态")
        if filter_active_products and product_type == "本品" and "在投状态" in lookup:
            if active_status.strip().lower() not in {x.lower() for x in ACTIVE_VALUES}:
                continue

        target = ScanTarget(
            country=country,
            product_type=product_type,
            portfolio_brand=_value(row, lookup, "Portfolio/品牌"),
            product=_value(row, lookup, "产品"),
            asin=asin,
            configuration=_value(row, lookup, "配置"),
            color=_value(row, lookup, "颜色"),
            url=url,
            active_status=active_status,
            remark=_value(row, lookup, "备注"),
        )
        key = (country, asin)
        if key not in merged:
            merged[key] = target
        else:
            existing = merged[key]
            # Fill blank metadata from duplicates, but still scan the ASIN only once per country.
            for attr in ("portfolio_brand", "product", "configuration", "color", "url", "active_status", "remark"):
                if not getattr(existing, attr) and getattr(target, attr):
                    setattr(existing, attr, getattr(target, attr))
    return list(merged.values())


def _style_sheet(ws, widths: dict[int, int] | None = None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if widths:
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _own_row(result: ScanResult):
    t, s = result.target, result.snapshot
    note = t.remark
    if result.scan_status != "OK":
        tech = f"{result.scan_status}: {result.error_reason}".strip(": ")
        note = f"{note}；{tech}".strip("；") if note else tech
    return [
        date.today(), t.country, t.portfolio_brand, t.product, t.asin, t.url, s.product_name,
        t.configuration, t.color, s.price_raw, s.price_value, s.list_price, s.buybox_seller,
        s.stock_text, s.delivery_text, s.delivery_over_10_days, s.rating, s.reviews,
        s.bsr_primary, s.bsr_secondary, result.anomaly, s.deal_tag, s.amazon_choice, note,
        s.purchase_box_status,
    ]


def _comp_row(result: ScanResult):
    t, s = result.target, result.snapshot
    return [
        date.today(), t.country, t.portfolio_brand, t.product, t.asin, t.url, s.product_name,
        t.configuration, t.color, s.price_raw, s.price_value, s.list_price, s.buybox_seller,
        s.stock_text, s.delivery_text, s.delivery_over_10_days, s.rating, s.reviews,
        s.bsr_primary, s.bsr_secondary,
    ]


def write_results(path: str | Path, results: list[ScanResult], started_at: datetime | None = None):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_own = wb.active
    ws_own.title = "本品扫查明细"
    ws_comp = wb.create_sheet("竞品扫查明细")
    ws_tech = wb.create_sheet("运行日志")
    ws_summary = wb.create_sheet("运行摘要")

    ws_own.append(OWN_HEADERS)
    ws_comp.append(COMP_HEADERS)
    ws_tech.append(TECH_HEADERS)

    for result in results:
        if result.target.product_type == "竞品":
            ws_comp.append(_comp_row(result))
        else:
            ws_own.append(_own_row(result))
        ws_tech.append([
            datetime.now(), result.target.country, result.target.product_type, result.target.asin,
            result.target.url, result.scan_status, result.error_reason, result.attempts,
            result.snapshot.page_state, result.snapshot.stock_status, result.snapshot.purchase_box_status,
            result.snapshot.buybox_seller, result.snapshot.ships_from, result.snapshot.page_title, result.debug_screenshot,
        ])

    for ws in (ws_own, ws_comp):
        for cell in ws["A"][1:]:
            cell.number_format = "yyyy-mm-dd"
    for cell in ws_tech["A"][1:]:
        cell.number_format = "yyyy-mm-dd hh:mm:ss"

    _style_sheet(ws_own, {1: 12, 2: 9, 3: 28, 4: 22, 5: 14, 6: 42, 7: 42, 8: 14, 9: 14, 10: 16, 11: 12, 12: 16, 13: 20, 14: 24, 15: 36, 21: 22, 24: 30, 25: 18})
    _style_sheet(ws_comp, {1: 12, 2: 9, 3: 18, 4: 22, 5: 14, 6: 42, 7: 42, 8: 14, 9: 14, 10: 16, 11: 12, 12: 16, 13: 20, 14: 24, 15: 36})
    _style_sheet(ws_tech, {1: 20, 2: 9, 3: 10, 4: 14, 5: 42, 6: 12, 7: 35, 8: 10, 9: 20, 10: 18, 11: 20, 12: 50, 13: 50})

    total = len(results)
    own = sum(r.target.product_type != "竞品" for r in results)
    comp = total - own
    failed = sum(r.scan_status == "FAILED" for r in results)
    partial = sum(r.scan_status == "PARTIAL" for r in results)
    anomalies = sum(bool(r.anomaly) for r in results if r.target.product_type != "竞品")
    summary_rows = [
        ["指标", "值"],
        ["开始时间", started_at or datetime.now()],
        ["结束时间", datetime.now()],
        ["总扫描ASIN", total],
        ["本品ASIN", own],
        ["竞品ASIN", comp],
        ["FAILED", failed],
        ["PARTIAL", partial],
        ["本品异常数", anomalies],
    ]
    for row in summary_rows:
        ws_summary.append(row)
    _style_sheet(ws_summary, {1: 22, 2: 28})
    ws_summary["B2"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws_summary["B3"].number_format = "yyyy-mm-dd hh:mm:ss"

    wb.save(path)


def create_target_template(path: str | Path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "scan_targets"
    headers = ["国家", "类型", "Portfolio/品牌", "产品", "ASIN", "配置", "颜色", "URL", "在投状态", "备注"]
    ws.append(headers)
    _style_sheet(ws, {1: 10, 2: 10, 3: 28, 4: 24, 5: 14, 6: 16, 7: 16, 8: 45, 9: 12, 10: 32})
    wb.save(path)
